from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def parse_csv_bytes(file_bytes: bytes, encoding: str = "utf-8") -> list[dict[str, Any]]:
    text = file_bytes.decode(encoding, errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [{key: _clean_cell(value) for key, value in row.items()} for row in reader]


def parse_json_bytes(file_bytes: bytes, encoding: str = "utf-8") -> list[dict[str, Any]]:
    payload = json.loads(file_bytes.decode(encoding, errors="replace"))
    if isinstance(payload, list):
        return [dict(item) for item in payload]
    if isinstance(payload, dict):
        rows = payload.get("rows") or payload.get("items") or payload.get("data")
        if isinstance(rows, list):
            return [dict(item) for item in rows]
        return [dict(payload)]
    raise ValueError("Unsupported JSON dataset format")


def parse_excel_bytes(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else f"column_{index + 1}" for index, cell in enumerate(rows[0])]
    output: list[dict[str, Any]] = []
    for row in rows[1:]:
        output.append({headers[index]: _clean_cell(value) for index, value in enumerate(row)})
    return output


def parse_dataset_bytes(file_name: str, file_bytes: bytes) -> list[dict[str, Any]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return parse_csv_bytes(file_bytes)
    if suffix in {".json"}:
        return parse_json_bytes(file_bytes)
    if suffix in {".xls", ".xlsx"}:
        return parse_excel_bytes(file_bytes)
    raise ValueError(f"Unsupported file type: {suffix or 'unknown'}")
